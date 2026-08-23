#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YENI LIG ICIN EXCEL URETME - Ingiltere Premier Lig dosyalarini SABLON
olarak kopyalar (KENDİLERİNE HİÇBİR SEKILDE DOKUNMAZ), her kopyayi
tamamen temizleyip yeni ligin GERCEK, DOGRULANMIS verisiyle doldurur.

SU AN SADECE 20 TAKIMLI, SEZON BOYUNCA TAKIM SAYISI DEGISMEYEN ligler
icin (Italya Serie A, Brezilya Serie A, Brezilya Serie B). 18/22 takimli
ya da sezon icinde takim sayisi degisen ligler (Bundesliga, Eredivisie,
Ligue 1/2, Italya Serie B) bu script'te YOK - sablon blok boyutu farkli
gerektiriyor, ayri bir asama.

NE YAPAR (her lig icin, her "N. HAFTA.xlsx" icin):
  1) "Ingiltere Premier Lig N. HAFTA.xlsx" dosyasini SADECE OKUR, KOPYALAR
     - orijinal dosyaya asla yazma islemi yapilmaz.
  2) Kopyada 25 sezon blogunu (26/27 + 24 tarihi sezon) bulur, HER
     BLOGUN eski Premier Lig verisini (takim adlari, puan durumu,
     mac verisi) TAMAMEN TEMIZLER.
  3) Blok basliklarini ve sezon etiketlerini YENI LIGE gore yeniden
     yazar (26/27 blogu HER ZAMAN bos birakilir - dokunulmaz).
  4) Elimizdeki dogrulanmis mac verisinden PUAN DURUMUNU kendimiz
     hesaplayip (Toplam + Icerde + Disarda) yaziyoruz - 38 dosyanin
     HER birinde o dosyanin kendi haftasi icin.
  5) MAClAR blogunu (AL:BD), Premier Lig'de kullanilan AYNI +1 ofset
     kuraliyla dolduruyor (N.HAFTA.xlsx -> Matchday N+1), 38. HAFTA'ya
     dokunmuyor.
  6) FORMULLERE (AF:AK, BE:JZ) KESINLIKLE DOKUNULMAZ - sadece veri
     hucreleri degisir.

Veri kaynagi: bu script "data_world/<lig>_maclar_TUM.json" dosyalarini
okur (world_leagues_fetch.py'nin urettigi, zaten dogrulanmis veri).

Calistirma:
    python build_league_excel.py
"""

import json
import os
import re
import shutil
import sys
import time
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SOURCE_DIR = BASE_DIR          # "Ingiltere Premier Lig N. HAFTA.xlsx" burada
OUTPUT_DIR = BASE_DIR          # yeni dosyalar da buraya yazilir
DATA_WORLD_DIR = os.path.join(BASE_DIR, "data_world")

SOURCE_FILENAME_RE = re.compile(r"İngiltere Premier Lig\s+(\d{1,2})\s*\.\s*HAFTA\.xlsx$", re.IGNORECASE)
TITLE_RE = re.compile(r"^(\S+)\s")  # blok basligindaki ilk "kelime" = sezon etiketi

N_TEAM_ROWS = 20
N_MATCH_ROWS = 10
N_BLOCKS = 25          # 26/27 (bos) + 24 tarihi sezon
WEEKS = 38
BLOCK_SCAN_MAX_ROW = 700

# --- Sablon en fazla 20 takim/blok destekliyor - 20'den fazla takimli
# sezonlar (orn. Italya Serie B'nin 22/24 takimli eski sezonlari) otomatik
# olarak elenir (asagidaki valid_seasons kontrolu), 18 takimli sezonlar ise
# ayni blokta sadece 18 satir kullanilarak sorunsuz calisir. ---
LEAGUES = [
    {"name": "İtalya Serie A", "data_file": "seriea_it_maclar_TUM.json", "season_kind": "range"},
    {"name": "Brezilya Serie A", "data_file": "brasileirao_a_maclar_TUM.json", "season_kind": "year"},
    {"name": "Brezilya Serie B", "data_file": "brasileirao_b_maclar_TUM.json", "season_kind": "year"},
    {"name": "İtalya Serie B", "data_file": "serieb_it_maclar_TUM.json", "season_kind": "range"},
    {"name": "Hollanda Eredivisie", "data_file": "eredivisie_maclar_TUM.json", "season_kind": "range"},
    {"name": "Fransa Ligue 1", "data_file": "ligue1_maclar_TUM.json", "season_kind": "range"},
    {"name": "Fransa Ligue 2", "data_file": "ligue2_maclar_TUM.json", "season_kind": "range"},
]

STANDINGS_TOTAL_COLS = ["E", "F", "G", "H", "I", "J", "K", "M"]        # O G B M A Y P (L bos) AV
STANDINGS_HOME_COLS = ["N", "O", "P", "Q", "R", "S", "T", "U"]          # O G B M A Y P AV (İCERDE)
STANDINGS_AWAY_COLS = ["V", "W", "X", "Y", "Z", "AA", "AB", "AC"]       # O G B M A Y P AV (DISARDA)

MATCH_COL_DATE = "AL"
MATCH_COL_MS = "AM"
MATCH_COL_HOME_RANK = "AN"
MATCH_COL_HOME = "AO"
MATCH_COL_FT = "AQ"
MATCH_COL_AWAY = "AS"
MATCH_COL_AWAY_RANK = "AT"
MATCH_COL_HT = "AU"


def log(msg):
    print(msg, flush=True)


# =============================================================================
# Veri yukleme + standart hesaplamalar (premier_test.py ile ayni mantik)
# =============================================================================

def load_league_matches(league):
    path = os.path.join(DATA_WORLD_DIR, league["data_file"])
    if not os.path.exists(path):
        raise RuntimeError(f"{path} bulunamadi - once world_leagues_fetch.py calistirilmali")
    with open(path, encoding="utf-8") as f:
        matches = json.load(f)
    return matches


def season_sort_key(season_label, season_kind):
    if season_kind == "year":
        return int(season_label)
    return int(season_label.split("/")[0])


def compute_standings_row(matches_upto_week, team):
    """Bir takimin (Toplam, Icerde, Disarda) istatistiklerini doner."""
    stats = {
        "total": {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0},
        "home": {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0},
        "away": {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0},
    }

    def apply(bucket, gf, ga):
        bucket["O"] += 1
        bucket["A"] += gf
        bucket["Y"] += ga
        if gf > ga:
            bucket["G"] += 1
            bucket["P"] += 3
        elif gf < ga:
            bucket["M"] += 1
        else:
            bucket["B"] += 1
            bucket["P"] += 1

    for m in matches_upto_week:
        if m["home"] == team:
            apply(stats["total"], m["home_goals"], m["away_goals"])
            apply(stats["home"], m["home_goals"], m["away_goals"])
        elif m["away"] == team:
            apply(stats["total"], m["away_goals"], m["home_goals"])
            apply(stats["away"], m["away_goals"], m["home_goals"])
    return stats


def compute_rank_table(season_matches, through_week):
    relevant = [m for m in season_matches if m["week"] <= through_week]
    table = defaultdict(lambda: {"P": 0, "A": 0, "Y": 0})
    for m in relevant:
        h, a = m["home"], m["away"]
        hg, ag = m["home_goals"], m["away_goals"]
        table[h]["A"] += hg; table[h]["Y"] += ag
        table[a]["A"] += ag; table[a]["Y"] += hg
        if hg > ag:
            table[h]["P"] += 3
        elif hg < ag:
            table[a]["P"] += 3
        else:
            table[h]["P"] += 1; table[a]["P"] += 1
    ranked = sorted(table.items(), key=lambda kv: (-kv[1]["P"], -(kv[1]["A"] - kv[1]["Y"]), -kv[1]["A"]))
    return {team: i + 1 for i, (team, _) in enumerate(ranked)}


# =============================================================================
# Excel yapisi (Premier Lig / La Liga sablonundan programatik olarak
# tespit edildi - tahmin edilmedi)
# =============================================================================

def find_source_files():
    found = {}
    for fn in os.listdir(SOURCE_DIR):
        if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
            continue
        m = SOURCE_FILENAME_RE.search(fn)
        if m:
            found[int(m.group(1))] = os.path.join(SOURCE_DIR, fn)
    return found


def discover_season_blocks(ws):
    blocks = []
    for r in range(1, BLOCK_SCAN_MAX_ROW + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip():
            blocks.append(r)
            if len(blocks) >= N_BLOCKS:
                break
    return blocks


def clear_block(ws, title_row, league_label_full):
    """Eski Premier Lig verisini TAMAMEN temizler: takim adlari, TUM
    puan durumu (toplam+icerde+disarda), sezon/hafta etiketleri. Formul
    zonuna (AF:AK, BE:JZ) VE mac blogunun kendisine (AL:BD) DOKUNMAZ -
    mac blogu ayri asamada, sadece veri varsa yazilir/temizlenir."""
    ws.cell(row=title_row, column=2).value = league_label_full  # B kolonu
    for i in range(N_TEAM_ROWS):
        r = title_row + 3 + i
        ws.cell(row=r, column=4).value = None  # D = TAKIM
        for col in STANDINGS_TOTAL_COLS + STANDINGS_HOME_COLS + STANDINGS_AWAY_COLS:
            ws[f"{col}{r}"] = None
        ws[f"AD{r}"] = None  # SEZON
        ws[f"AE{r}"] = None  # HAFTA
    for i in range(N_MATCH_ROWS):
        r = title_row + 3 + i
        for col in (MATCH_COL_DATE, MATCH_COL_MS, MATCH_COL_HOME_RANK, MATCH_COL_HOME,
                    MATCH_COL_FT, MATCH_COL_AWAY, MATCH_COL_AWAY_RANK, MATCH_COL_HT):
            ws[f"{col}{r}"] = None
            ws[f"{col}{r}"].fill = NO_FILL_PLACEHOLDER
            ws[f"{col}{r}"].font = DEFAULT_FONT_PLACEHOLDER


def write_standings_block(ws, title_row, season_label, week_n, teams_sorted, season_matches):
    relevant = [m for m in season_matches if m["week"] <= week_n]
    for i, team in enumerate(teams_sorted):
        r = title_row + 3 + i
        ws.cell(row=r, column=4).value = team
        ws[f"AD{r}"] = season_label
        ws[f"AE{r}"] = week_n

        stats = compute_standings_row(relevant, team)
        for label, cols in (("total", STANDINGS_TOTAL_COLS), ("home", STANDINGS_HOME_COLS), ("away", STANDINGS_AWAY_COLS)):
            s = stats[label]
            vals = [s["O"], s["G"], s["B"], s["M"], s["A"], s["Y"], s["P"], s["A"] - s["Y"]]
            for col, val in zip(cols, vals):
                ws[f"{col}{r}"] = val


def write_match_block(ws, title_row, target_matchday, season_matches, rank_table, wb_styles, expected_match_count):
    matches = [m for m in season_matches if m["week"] == target_matchday]
    if len(matches) != expected_match_count:
        return False, len(matches)
    for i, m in enumerate(matches):
        r = title_row + 3 + i
        if m.get("date_dm"):
            ws[f"{MATCH_COL_DATE}{r}"] = m["date_dm"]
        ws[f"{MATCH_COL_MS}{r}"] = "MS"
        ws[f"{MATCH_COL_HOME_RANK}{r}"] = rank_table.get(m["home"])
        ws[f"{MATCH_COL_HOME}{r}"] = m["home"]
        ws[f"{MATCH_COL_FT}{r}"] = f"{m['home_goals']} - {m['away_goals']}"
        ws[f"{MATCH_COL_AWAY}{r}"] = m["away"]
        ws[f"{MATCH_COL_AWAY_RANK}{r}"] = rank_table.get(m["away"])
        if m.get("ht_home_goals") is not None:
            ws[f"{MATCH_COL_HT}{r}"] = f"{m['ht_home_goals']} - {m['ht_away_goals']}"

        home_cell, away_cell = ws[f"{MATCH_COL_HOME}{r}"], ws[f"{MATCH_COL_AWAY}{r}"]
        if m["home_goals"] > m["away_goals"]:
            home_cell.fill, home_cell.font = wb_styles["winner_fill"], wb_styles["winner_font"]
            away_cell.fill, away_cell.font = wb_styles["no_fill"], wb_styles["default_font"]
        elif m["away_goals"] > m["home_goals"]:
            away_cell.fill, away_cell.font = wb_styles["winner_fill"], wb_styles["winner_font"]
            home_cell.fill, home_cell.font = wb_styles["no_fill"], wb_styles["default_font"]
        else:
            for c in (home_cell, away_cell):
                c.fill, c.font = wb_styles["no_fill"], wb_styles["default_font"]
    return True, len(matches)


NO_FILL_PLACEHOLDER = None
DEFAULT_FONT_PLACEHOLDER = None


def build_league(league, source_files, openpyxl_mod):
    from openpyxl.styles import PatternFill, Font
    global NO_FILL_PLACEHOLDER, DEFAULT_FONT_PLACEHOLDER
    NO_FILL_PLACEHOLDER = PatternFill(fill_type=None)
    DEFAULT_FONT_PLACEHOLDER = Font()
    wb_styles = {
        "winner_fill": PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid"),
        "winner_font": Font(color="FFFFFFFF", bold=True),
        "no_fill": NO_FILL_PLACEHOLDER,
        "default_font": DEFAULT_FONT_PLACEHOLDER,
    }

    log("=" * 70)
    log(league["name"])
    log("=" * 70)

    all_matches = load_league_matches(league)
    by_season = defaultdict(list)
    for m in all_matches:
        by_season[m["season"]].append(m)

    # sadece TAM (N takim x cift devre x N*(N-1) mac) dogrulanan sezonlar.
    # N en fazla 20 olabilir (sablonun blok basina 20 satir siniri) - daha
    # fazla takimli sezonlar (orn. Italya Serie B'nin 22/24 takimli eski
    # sezonlari) burada otomatik elenir, tahmini/yanlis veri yazilmaz.
    valid_seasons = []
    for season, ms in by_season.items():
        teams = set(m["home"] for m in ms) | set(m["away"] for m in ms)
        n_teams = len(teams)
        weeks = set(m["week"] for m in ms)
        n_expected_weeks = 2 * (n_teams - 1)
        expected_total = n_teams * (n_teams - 1)
        if 2 <= n_teams <= N_TEAM_ROWS and len(ms) == expected_total and weeks == set(range(1, n_expected_weeks + 1)):
            valid_seasons.append(season)
    valid_seasons.sort(key=lambda s: season_sort_key(s, league["season_kind"]), reverse=True)
    valid_seasons = valid_seasons[:24]  # sablon 24 tarihi blok destekliyor

    log(f"  Kullanilacak sezon sayisi: {len(valid_seasons)} -> {valid_seasons}")
    if not valid_seasons:
        log("  UYARI: gecerli sezon yok, bu lig atlaniyor.")
        return

    team_order = {}
    for season in valid_seasons:
        teams = sorted(set(m["home"] for m in by_season[season]) | set(m["away"] for m in by_season[season]))
        team_order[season] = teams

    newest_year = season_sort_key(valid_seasons[0], league["season_kind"]) + 1
    if league["season_kind"] == "year":
        future_season_label = str(newest_year)
    else:
        future_season_label = f"{newest_year % 100:02d}/{(newest_year + 1) % 100:02d}"

    files_written = 0
    for week_n in range(1, WEEKS + 1):
        src = source_files.get(week_n)
        if not src:
            log(f"  UYARI: kaynak {week_n}. HAFTA.xlsx bulunamadi, atlaniyor.")
            continue

        dst_name = f"{league['name']} {week_n}. HAFTA.xlsx"
        dst_path = os.path.join(OUTPUT_DIR, dst_name)

        copy_ok = False
        for attempt in range(5):
            try:
                shutil.copy2(src, dst_path)  # KAYNAK asla degistirilmez
                copy_ok = True
                break
            except PermissionError:
                log(f"  UYARI ({dst_name}): dosya kilitli (Excel'de acik ya da OneDrive senkronize ediyor olabilir), "
                    f"3 saniye sonra tekrar denenecek (deneme {attempt + 1}/5)...")
                time.sleep(3)
        if not copy_ok:
            log(f"  HATA ({dst_name}): 5 denemede de dosya acilamadi (kilitli kaldi), atlaniyor. "
                f"Excel'de acik degilse OneDrive'in senkronu bitirmesini bekleyip tekrar calistir.")
            continue

        wb = openpyxl_mod.load_workbook(dst_path, data_only=False)
        sheet_name = "SAYFA 1" if "SAYFA 1" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        block_rows = discover_season_blocks(ws)
        if len(block_rows) < N_BLOCKS:
            log(f"  HATA ({dst_name}): {len(block_rows)} blok bulundu, {N_BLOCKS} bekleniyordu. Dosya atlaniyor.")
            continue

        # blok 0: HER ZAMAN bos/gelecek sezon - sadece basligi yeniden yazip dokunmuyoruz
        future_row = block_rows[0]
        ws.cell(row=future_row, column=2).value = f"{future_season_label} {league['name']} {week_n}. HAFTA"

        match_written_total = 0
        for idx, season in enumerate(valid_seasons, start=1):
            title_row = block_rows[idx]
            season_matches = by_season[season]
            clear_block(ws, title_row, f"{season} {league['name']} {week_n}. HAFTA")
            write_standings_block(ws, title_row, season, week_n, team_order[season], season_matches)

            if week_n < WEEKS:  # 38. HAFTA'nin mac alanina KESINLIKLE dokunulmaz
                target_matchday = week_n + 1
                rank_table = compute_rank_table(season_matches, week_n)
                expected_match_count = len(team_order[season]) // 2
                ok, n = write_match_block(ws, title_row, target_matchday, season_matches, rank_table,
                                           wb_styles, expected_match_count)
                if ok:
                    match_written_total += n

        # kullanilmayan tarihi bloklari (valid_seasons'dan fazla olanlari) da temizle
        for extra_row in block_rows[1 + len(valid_seasons):N_BLOCKS]:
            clear_block(ws, extra_row, f"(veri yok) {league['name']} {week_n}. HAFTA")

        wb.save(dst_path)
        files_written += 1
        log(f"  -> {dst_name}: {len(valid_seasons)} sezon puan durumu, {match_written_total} mac yazildi")

    log(f"  TOPLAM: {files_written} / {WEEKS} dosya uretildi.")


def main():
    try:
        import openpyxl
    except ImportError:
        log("HATA: openpyxl kurulu degil. Once calistir: pip install openpyxl")
        sys.exit(1)

    source_files = find_source_files()
    if not source_files:
        log(f"HATA: {SOURCE_DIR} icinde 'İngiltere Premier Lig N. HAFTA.xlsx' dosyalari bulunamadi.")
        log("Bu dosyalar SABLON olarak kullanilir - kendileri degistirilmez.")
        sys.exit(1)
    log(f"Kaynak sablon: {len(source_files)} 'İngiltere Premier Lig N. HAFTA.xlsx' dosyasi bulundu.")

    for league in LEAGUES:
        try:
            build_league(league, source_files, openpyxl)
        except RuntimeError as e:
            log(f"HATA ({league['name']}): {e}")

    log("=" * 70)
    log("BITTI. Ingiltere Premier Lig dosyalarina HICBIR sekilde dokunulmadi.")


if __name__ == "__main__":
    main()
