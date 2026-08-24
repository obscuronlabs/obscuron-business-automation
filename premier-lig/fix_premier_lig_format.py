#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ingiltere Premier Lig N. HAFTA.xlsx dosyalarindaki SADECE GORSEL iki
eksikligi duzeltir - kullanicinin ACIK ONAYIYLA:
  1) HAFTA (AE) sutununda bazi satirlarda "8.0"/"5.0"/"2.0" gibi
     ondalikli gorunen hafta numaralarini "8"/"5"/"2" yapar - SADECE
     sayi formati degisir, deger (8, 5, 2...) ayniyla kalir.
  2) 11-20. takim satirlarinda eksik olan kenarlik (border) cizgilerini
     1. siradaki (rank 1) satirin kenarligini kopyalayarak tamamlar -
     tum 25 blokta (26/27 dahil).

HICBIR FORMULE (AF:AK, BE:JZ), PUAN DURUMU VERISINE, MAC VERISINE
DOKUNULMAZ - sadece AE hucrelerinin sayi formati ve kenarlik bilgisi
degisir.

ONCE TUM DOSYALARIN YEDEGINI ALIR (_BACKUP_FORMAT_DUZELTME/ altina,
zaten yedeklenmis dosyalar tekrar yedeklenmez).

Calistirma:
    python fix_premier_lig_format.py
"""
import os
import re
import shutil
import sys
import time
from copy import copy

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BACKUP_DIR = os.path.join(BASE_DIR, "_BACKUP_FORMAT_DUZELTME")

FILENAME_RE = re.compile(r"İngiltere Premier Lig\s+(\d{1,2})\s*\.\s*HAFTA\.xlsx$", re.IGNORECASE)
N_TEAM_ROWS = 20
N_BLOCKS = 25
BLOCK_SCAN_MAX_ROW = 700

STANDINGS_TOTAL_COLS = ["E", "F", "G", "H", "I", "J", "K", "M"]
STANDINGS_HOME_COLS = ["N", "O", "P", "Q", "R", "S", "T", "U"]
STANDINGS_AWAY_COLS = ["V", "W", "X", "Y", "Z", "AA", "AB", "AC"]
# Satirin TAMAMI (A'dan AE'ye kadar, C=sira numarasi dahil) - onceki
# surumde sadece belirli sutunlar listelenmisti ve C (sira no) atlanmisti,
# bu yuzden 11-20. satirlarda kenarlik hala eksik gorunuyordu.
BORDER_REFERENCE_COLS = ["A", "B", "C", "D"] + STANDINGS_TOTAL_COLS + ["L"] + STANDINGS_HOME_COLS + STANDINGS_AWAY_COLS + ["AD", "AE"]


def log(msg):
    print(msg, flush=True)


def find_files():
    found = {}
    for fn in os.listdir(BASE_DIR):
        if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
            continue
        m = FILENAME_RE.search(fn)
        if m:
            found[int(m.group(1))] = os.path.join(BASE_DIR, fn)
    return found


def discover_season_blocks(ws):
    blocks = []
    for r in range(1, BLOCK_SCAN_MAX_ROW + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip():
            blocks.append(r)
            if len(blocks) >= N_BLOCKS:
                break
    return blocks


def fix_block(ws, title_row):
    ref_row = title_row + 3
    # 1) AE (HAFTA) sayi formatini duzelt - DEGERI DEGISTIRMEDEN
    for i in range(N_TEAM_ROWS):
        r = title_row + 3 + i
        cell = ws[f"AE{r}"]
        if cell.value is not None:
            cell.number_format = "0"
    # 2) kenarliklari 1. siradan (rank 1) kopyala
    for col in BORDER_REFERENCE_COLS:
        ref_border = ws[f"{col}{ref_row}"].border
        for i in range(1, N_TEAM_ROWS):
            r = title_row + 3 + i
            ws[f"{col}{r}"].border = copy(ref_border)


def main():
    try:
        import openpyxl
    except ImportError:
        log("HATA: openpyxl kurulu degil. Once calistir: pip install openpyxl")
        sys.exit(1)

    files = find_files()
    if not files:
        log("HATA: 'İngiltere Premier Lig N. HAFTA.xlsx' dosyalari bulunamadi.")
        sys.exit(1)
    log(f"{len(files)} dosya bulundu.\n")

    os.makedirs(BACKUP_DIR, exist_ok=True)
    log(f"Yedekleniyor -> {BACKUP_DIR}")
    for week_n, path in sorted(files.items()):
        backup_path = os.path.join(BACKUP_DIR, os.path.basename(path))
        if not os.path.exists(backup_path):
            shutil.copy2(path, backup_path)
    log("Yedekleme tamam.\n")

    fixed = 0
    for week_n, path in sorted(files.items()):
        fname = os.path.basename(path)

        wb = None
        for attempt in range(5):
            try:
                wb = openpyxl.load_workbook(path, data_only=False)
                break
            except PermissionError:
                log(f"  UYARI ({fname}): dosya kilitli (Excel'de acik olabilir), 3 saniye sonra tekrar denenecek...")
                time.sleep(3)
        if wb is None:
            log(f"  HATA ({fname}): dosya acilamadi (kilitli kaldi), atlaniyor.")
            continue

        sheet_name = "SAYFA 1" if "SAYFA 1" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        block_rows = discover_season_blocks(ws)
        if len(block_rows) < N_BLOCKS:
            log(f"  UYARI ({fname}): {len(block_rows)} blok bulundu, {N_BLOCKS} bekleniyordu. Atlaniyor.")
            continue

        for title_row in block_rows:
            fix_block(ws, title_row)

        save_ok = False
        for attempt in range(5):
            try:
                wb.save(path)
                save_ok = True
                break
            except PermissionError:
                log(f"  UYARI ({fname}): kaydedilirken kilitli (Excel'de acik olabilir), 3 saniye sonra tekrar denenecek...")
                time.sleep(3)
        if not save_ok:
            log(f"  HATA ({fname}): kaydedilemedi (kilitli kaldi), atlaniyor.")
            continue

        fixed += 1
        log(f"  -> {fname}: {len(block_rows)} blok duzeltildi")

    log(f"\nBITTI. {fixed}/{len(files)} dosya duzeltildi.")
    log("Sadece AE (HAFTA) sayi formati ve kenarliklar degisti - hicbir veri/formul degismedi.")
    log(f"Orijinal halleri {BACKUP_DIR} icinde yedekli duruyor.")


if __name__ == "__main__":
    main()
