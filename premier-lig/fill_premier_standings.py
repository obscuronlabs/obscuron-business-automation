#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ingiltere Premier Lig dosyalarindaki ICERDE/DISARDA (N:AC) hucrelerini,
Mackolik'ten cekilen data_world/ingiltere_premier_maclar_TUM.json
verisinden hesaplayip doldurur.

GUVENLIK KURALI: SADECE BOS (None) hucreler doldurulur. Zaten elle
girilmis bir deger varsa ASLA UZERINE YAZILMAZ, dokunulmaz bile.
Bu yuzden yanlislikla mevcut dogru veriyi bozma riski yoktur - sadece
eksik olan kisim tamamlanir.

TOPLAM (E:M), formuller (AF:AK, BE:JZ), MAClAR (AL:AU), sezon/hafta
etiketleri (AD/AE) - HICBIRINE dokunulmaz.

Once TUM dosyalarin yedegi alinir.

Calistirma:
    python fill_premier_standings.py
"""
import json
import os
import re
import shutil
import sys
import time

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(BASE_DIR, "data_world", "ingiltere_premier_maclar_TUM.json")
BACKUP_DIR = os.path.join(BASE_DIR, "_BACKUP_ICERDE_DISARDA")

import premier_test as pt

# premier_test.py'nin kendi find_excel_files'i "N. HAFTA.xlsx" ile biten
# HER dosyayi eslestiriyor (klasordeki diger 8 ligin dosyalari dahil) -
# burada SADECE Ingiltere Premier Lig dosyalarini istiyoruz.
PREMIER_FILENAME_RE = re.compile(r"^İngiltere Premier Lig\s+(\d{1,2})\s*\.\s*HAFTA\.xlsx$", re.IGNORECASE)


def find_premier_files(directory):
    found = {}
    for fn in os.listdir(directory):
        if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
            continue
        m = PREMIER_FILENAME_RE.match(fn)
        if m:
            found[int(m.group(1))] = os.path.join(directory, fn)
    return found


STANDINGS_HOME_COLS = ["N", "O", "P", "Q", "R", "S", "T", "U"]
STANDINGS_AWAY_COLS = ["V", "W", "X", "Y", "Z", "AA", "AB", "AC"]
STAT_LABELS = ["O", "G", "B", "M", "A", "Y", "P", "AV"]
N_TEAM_ROWS = 20


def log(msg):
    print(msg, flush=True)


def compute_standings_row(matches_upto_week, team):
    stats = {
        "home": {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0},
        "away": {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0},
    }

    def apply(bucket, gf, ga):
        bucket["O"] += 1
        bucket["A"] += gf
        bucket["Y"] += ga
        if gf > ga:
            bucket["G"] += 1
            bucket["P"] += 3
        elif gf < ga:
            bucket["M"] += 1
        else:
            bucket["B"] += 1
            bucket["P"] += 1

    for m in matches_upto_week:
        if m["home"] == team:
            apply(stats["home"], m["home_goals"], m["away_goals"])
        elif m["away"] == team:
            apply(stats["away"], m["away_goals"], m["home_goals"])
    return stats


def main():
    try:
        import openpyxl
    except ImportError:
        log("HATA: openpyxl kurulu degil. Once calistir: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(DATA_PATH):
        log(f"HATA: {DATA_PATH} bulunamadi - once world_leagues_fetch.py calistirilmali.")
        sys.exit(1)

    with open(DATA_PATH, encoding="utf-8") as f:
        all_matches = json.load(f)
    by_season = {}
    for m in all_matches:
        by_season.setdefault(m["season"], []).append(m)
    log(f"data_world'den {len(all_matches)} mac, {len(by_season)} sezon yuklendi.\n")

    files = find_premier_files(BASE_DIR)
    if not files:
        log("HATA: 'İngiltere Premier Lig N. HAFTA.xlsx' dosyalari bulunamadi.")
        sys.exit(1)
    log(f"{len(files)} dosya bulundu.\n")

    os.makedirs(BACKUP_DIR, exist_ok=True)
    log(f"Yedekleniyor -> {BACKUP_DIR}")
    for week_n, path in sorted(files.items()):
        backup_path = os.path.join(BACKUP_DIR, os.path.basename(path))
        if not os.path.exists(backup_path):
            shutil.copy2(path, backup_path)
    log("Yedekleme tamam.\n")

    total_filled = 0
    total_files_changed = 0

    for week_n, path in sorted(files.items()):
        fname = os.path.basename(path)

        wb = None
        for attempt in range(5):
            try:
                wb = openpyxl.load_workbook(path, data_only=False)
                break
            except PermissionError:
                log(f"  UYARI ({fname}): dosya kilitli, 3 saniye sonra tekrar denenecek...")
                time.sleep(3)
        if wb is None:
            log(f"  HATA ({fname}): dosya acilamadi (kilitli kaldi), atlaniyor.")
            continue

        sheet_name = "SAYFA 1" if "SAYFA 1" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        blocks = pt.discover_season_blocks(ws)
        file_filled = 0

        for title_row, season_label in blocks:
            season_matches = by_season.get(season_label)
            if not season_matches:
                continue  # bu sezon Mackolik verisinde yok, atla - dokunma

            for i in range(N_TEAM_ROWS):
                r = title_row + 3 + i
                team = ws.cell(row=r, column=4).value
                if not team:
                    continue
                week_val = ws[f"AE{r}"].value
                if week_val is None:
                    continue
                through_week = int(week_val)

                relevant = [mm for mm in season_matches if mm["week"] <= through_week]
                computed = compute_standings_row(relevant, team)

                for label_set, cols in (("home", STANDINGS_HOME_COLS), ("away", STANDINGS_AWAY_COLS)):
                    s = computed[label_set]
                    vals = [s["O"], s["G"], s["B"], s["M"], s["A"], s["Y"], s["P"], s["A"] - s["Y"]]
                    for col, val in zip(cols, vals):
                        cell = ws[f"{col}{r}"]
                        if cell.value is None:  # SADECE BOS HUCRE - guvenlik kurali
                            cell.value = val
                            file_filled += 1

        if file_filled > 0:
            save_ok = False
            for attempt in range(5):
                try:
                    wb.save(path)
                    save_ok = True
                    break
                except PermissionError:
                    log(f"  UYARI ({fname}): kaydedilirken kilitli, 3 saniye sonra tekrar denenecek...")
                    time.sleep(3)
            if not save_ok:
                log(f"  HATA ({fname}): kaydedilemedi (kilitli kaldi), bu dosya ATLANDI.")
                continue
            total_files_changed += 1
            total_filled += file_filled
            log(f"  -> {fname}: {file_filled} bos hucre dolduruldu")
        else:
            log(f"  -> {fname}: doldurulacak bos hucre yok, dosya degistirilmedi")

    log(f"\nBITTI. {total_files_changed}/{len(files)} dosya guncellendi, toplam {total_filled} hucre dolduruldu.")
    log(f"Orijinal halleri {BACKUP_DIR} icinde yedekli duruyor.")
    log("SADECE bos hucreler dolduruldu - zaten dolu olan hicbir hucreye dokunulmadi.")


if __name__ == "__main__":
    main()
