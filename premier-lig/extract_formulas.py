#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bir "N. HAFTA.xlsx" dosyasindaki formul bolgesini (AF:AK ve BE:JZ) OKUR,
hicbir sey YAZMAZ/DEGISTIRMEZ. Amac: Poisson tahmin panelinin formul
mantigini incelemek.

Sadece TEK BIR bloktan (ilk gecerli sezon blogu) ornek alir - hepsi ayni
formul kalibinin farkli satirlara kopyalanmis hali oldugundan bir blok
yeterli.

Calistirma:
    python extract_formulas.py "İngiltere Premier Lig 10. HAFTA.xlsx"
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

N_TEAM_ROWS = 20
BLOCK_SCAN_MAX_ROW = 700


def discover_season_blocks(ws, limit=2):
    blocks = []
    for r in range(1, BLOCK_SCAN_MAX_ROW + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip():
            blocks.append(r)
            if len(blocks) >= limit:
                break
    return blocks


def dump_range(ws, title_row, col_start, col_end, label):
    c1 = column_index_from_string(col_start)
    c2 = column_index_from_string(col_end)
    print(f"\n{'=' * 70}\n{label} ({col_start}:{col_end}) - blok basligi satiri {title_row}\n{'=' * 70}")
    # baslik satirlarini da goster (formul sutun basliklarini anlamak icin)
    for r in range(title_row, title_row + 3):
        row_vals = []
        for c in range(c1, c2 + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_vals.append(f"{get_column_letter(c)}{r}={v!r}")
        if row_vals:
            print(f"  [baslik satiri {r}]: " + " | ".join(row_vals))

    # ilk 3 takim satirinin (rank 1-3) formullerini goster - hepsi ayni
    # kalip oldugundan 3 tanesi yeterli, tekrar tekrar 20 satir basmaya gerek yok
    for i in range(3):
        r = title_row + 3 + i
        print(f"\n  --- satir {r} (rank {i + 1}) ---")
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                print(f"    {get_column_letter(c)}{r}: {cell.value!r}")


def scan_headers(ws, title_row, col_start, col_end):
    """Sadece baslik satirlarini (title_row..title_row+2) tarar - genis bir
    aralikta NEYIN NEREDE oldugunu hizlica haritalamak icin. Formul metnini
    basmaz, cok genis araliklarda (orn. CE:JZ, ~200 sutun) once bunu
    calistirip sonra ilginc kismi hedefli cekmek icin kullan."""
    c1 = column_index_from_string(col_start)
    c2 = column_index_from_string(col_end)
    print(f"\n{'=' * 70}\nBASLIK TARAMASI ({col_start}:{col_end}) - blok basligi satiri {title_row}\n{'=' * 70}")
    for r in range(title_row, title_row + 3):
        for c in range(c1, c2 + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                print(f"  {get_column_letter(c)}{r} = {v!r}")


def main():
    if len(sys.argv) < 2:
        print("Kullanim: python extract_formulas.py \"İngiltere Premier Lig N. HAFTA.xlsx\" [BASLA] [BITIR] [--headers-only]")
        sys.exit(1)
    headers_only = "--headers-only" in sys.argv
    positional = [a for a in sys.argv[1:] if a != "--headers-only"]
    path = positional[0]
    custom_start = positional[1] if len(positional) > 1 else None
    custom_end = positional[2] if len(positional) > 2 else None

    # data_only=False -> formul METNINI okur (hesaplanmis DEGERI degil)
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_name = "SAYFA 1" if "SAYFA 1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    blocks = discover_season_blocks(ws, limit=2)
    if not blocks:
        print("HATA: blok bulunamadi.")
        sys.exit(1)

    # blok 0 = bos gelecek sezon (formul olsa da veri yok), blok 1 = dolu
    # tarihi sezon - formulleri asil ANLAMLI degerlerle gormek icin blok 1'i kullan
    title_row = blocks[1] if len(blocks) > 1 else blocks[0]

    if custom_start and custom_end:
        if headers_only:
            scan_headers(ws, title_row, custom_start, custom_end)
        else:
            dump_range(ws, title_row, custom_start, custom_end, f"FORMUL BOLGESI ({custom_start}:{custom_end})")
        return

    dump_range(ws, title_row, "AF", "AK", "FORMUL BOLGESI 1 (AF:AK)")
    dump_range(ws, title_row, "BE", "BL", "FORMUL BOLGESI 2 (BE:BL - Poisson panelinin ILK kismi)")
    dump_range(ws, title_row, "BM", "BU", "FORMUL BOLGESI 2 devam (BM:BU)")
    dump_range(ws, title_row, "BV", "CD", "FORMUL BOLGESI 2 devam (BV:CD)")

    print(f"\n\nNOT: Panel BE:JZ'ye kadar uzaniyor, yer kaplamasin diye once CD sutununa kadar cikti aldik.")
    print("Daha ilerisi (CE:JZ) icin: python extract_formulas.py \"<dosya>\" CE JZ --headers-only")


if __name__ == "__main__":
    main()
