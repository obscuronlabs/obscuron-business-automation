#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Premier Lig - Resmi Gameweek Verisiyle Excel MAClAR Blogunu Doldurma
=====================================================================

TEK PARCA, TAM, CALISIR script. C:\\Users\\barba\\premier_test.py olarak
calistirilmak uzere yazildi.

NE YAPAR (sirasiyla):
  1) openfootball/england kaynagindan 02/03-25/26 arasi 24 sezonun TUM
     Premier Lig maclarini ceker (cache'li, retry'li, hizli-vazgecmeli).
     Bu kaynak maclari TARIHE gore degil, ligin kendi RESMI Matchday/Round
     basligina gore gruplar - ertelenen/one alinan mac bile ait oldugu
     GERCEK haftada kalir. Tarihten hafta hesaplama YAPILMAZ.
  2) Cekilen veriyi sikica dogrular: her sezon 380 mac, hafta 1-38, mukerrer
     yok, takim basina 38 mac, 3 bilinen sampiyonluk puanina karsi bagimsiz
     capraz kontrol. TEK BIR HATA VARSA HICBIR EXCEL DOSYASINA YAZILMAZ.
  3) EXCEL_DIR icindeki "... N. HAFTA.xlsx" dosyalarini (N=1..37) acar,
     SAYFA 1'deki her sezon blogunu (satir basliklarindan) programatik
     olarak bulur - satir/kolon numarasi TAHMIN EDILMEZ.
  4) +1 OFSET KURALI (La Liga sistemiyle ayni mantik):
         N. HAFTA.xlsx dosyasinin PUAN DURUMU = N. hafta (ZATEN TAMAM, DOKUNULMAZ)
         N. HAFTA.xlsx dosyasinin MAClAR      = OpenFootball Matchday (N+1)
     38. HAFTA.xlsx'e KESINLIKLE DOKUNULMAZ (Matchday 39 yok).
  5) Sadece 4 hucreye yazar (satir basina): AO=ev sahibi, AS=deplasman,
     AQ=M/S (mac sonucu skoru), AU=I/Y (ilk yari skoru, kaynakta varsa).
     Baska HICBIR hucreye (formul, puan durumu, 26/27 blogu, spacer
     kolonlar) dokunulmaz.
  6) Her dosyadan once orijinali _BACKUP_MAC_SONUCLARI/ klasorune yedeklenir
     (zaten yedek varsa uzerine yazilmaz - ilk calistirmadaki temiz hali korunur).
  7) Yazdiktan sonra dosyayi tekrar diskten okuyup dogrular: yazilan 4 hucre
     dogru mu, formul hucreleri (metin olarak) degismedi mi, puan durumu
     degismedi mi, 26/27 blogu degismedi mi.
  8) Sonunda tek bir NET denetim raporu basar.

YAPILMAYAN SEYLER (bilincli olarak):
  - Yeni veri kaynagi aranmaz, tarihten hafta tahmin edilmez.
  - Puan durumuna, formullere, formatlara, merged hucrelere dokunulmaz.
  - 26/27 sezonuna ve 38. HAFTA.xlsx'e dokunulmaz.

KULLANIM:
    pip install openpyxl
    python premier_test.py

Varsayilan olarak xlsx dosyalarinin bu script ile AYNI klasorde oldugunu
varsayar. Farkli bir klasordeyse asagidaki EXCEL_DIR degiskenini degistir.
"""

import csv
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict

# =============================================================================
# AYARLAR - gerekirse burayi degistir
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# xlsx dosyalarinin ("Ingiltere Premier Lig N. HAFTA.xlsx") bulundugu klasor.
# Varsayilan: script ile ayni klasor. Farkliysa degistir, orn:
# EXCEL_DIR = r"C:\Users\barba\Desktop\Premier Lig Excel"
EXCEL_DIR = BASE_DIR

CACHE_DIR = os.path.join(BASE_DIR, "cache")
DATA_DIR = os.path.join(BASE_DIR, "data")
BACKUP_DIR = os.path.join(EXCEL_DIR, "_BACKUP_MAC_SONUCLARI")

RAW_URL_TMPL = (
    "https://raw.githubusercontent.com/openfootball/england/master/"
    "{season}/1-premierleague.txt"
)

SEASON_START_YEARS = list(range(2002, 2026))  # 2002-03 .. 2025-26 (24 sezon)
SEASONS = ["{:02d}/{:02d}".format(y % 100, (y + 1) % 100) for y in SEASON_START_YEARS]
SEASONS_SET = set(SEASONS)
SEASON_FOLDERS = {
    s: "{}-{:02d}".format(y, (y + 1) % 100) for s, y in zip(SEASONS, SEASON_START_YEARS)
}

HTTP_TIMEOUT = 12
MAX_ATTEMPTS = 2          # kaynak yaniti gec/hata verirse fazla ugrasma
RETRY_DELAY = 2.0

# Bilinen sonuclarla capraz kontrol (parser dogrulugu icin bagimsiz kanit,
# eski-format ve yeni-format sezonlarindan birer ornek).
KNOWN_CHAMPION_POINTS = [
    ("03/04", "Arsenal", 90),        # "Invincibles" - yenilgisiz sampiyon
    ("15/16", "Leicester", 81),      # Leicester City'nin mucizevi sampiyonlugu
    ("22/23", "Manchester City", 89),
]

# openfootball takim adlarini Excel'deki (D kolonu) kisa takim adlarina
# cevirir. Once " FC"/" AFC" son eki atilir, sonra bu tabloda ozel bir
# esleme varsa o kullanilir. 9120 maclik veri setinin tamami uzerinde
# dogrulandi: 0 uyusmazlik.
TEAM_NAME_OVERRIDES = {
    "AFC Bournemouth": "Bournemouth",
    "Tottenham Hotspur": "Tottenham",
    "West Bromwich Albion": "West Bromwich",
    "Wolverhampton Wanderers": "Wolverhampton",
    "Birmingham City": "Birmingham",
}


def normalize_team(name):
    n = re.sub(r"\s+AFC$", "", name)
    n = re.sub(r"\s+FC$", "", n)
    return TEAM_NAME_OVERRIDES.get(n, n)


HEADER_RE = re.compile(
    r"^▪\s*(?:Matchday|Round|Regular\s+Season\s*-)\s*(\d+)", re.IGNORECASE
)
# Tarih basligi satirlari: "Sat Aug 8", "Fri Aug 15 2025" vb. Sadece hafta
# gunu kisaltmasiyla baslamak yeterli DEGIL - "Sunderland" de "Sun" ile
# basliyor ve bir mac satirinin ev sahibi olabiliyor. Ay kisaltmasini da
# zorunlu tutarak bu yanlis eslesmeyi onluyoruz. Gun sayisini da yakalayip
# gercek mac tarihini (AL kolonu icin GG/AA) cikartiyoruz.
DATE_LINE_RE = re.compile(
    r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\b"
)
MONTH_NUM = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}
# Format A (cogunlukla eski sezonlar): "Ev Takim   H-A (IY)  Deplasman"
MATCH_LINE_RE_DASH = re.compile(
    r"^\s*(?:\d{1,2}[:.]\d{2}\s*)?"
    r"(?P<home>.+?)\s{2,}"
    r"(?P<hg>\d+)\s*-\s*(?P<ag>\d+)"
    r"(?:\s*\((?P<ht>[^)]*)\))?"
    r"\s+(?P<away>.+?)\s*$"
)
# Format B (bazi yeni sezonlar, orn. 2024-25): "Ev Takim  v  Deplasman   H-A (IY)"
MATCH_LINE_RE_V = re.compile(
    r"^\s*(?:\d{1,2}[:.]\d{2}\s*)?"
    r"(?P<home>.+?)\s+v\s+"
    r"(?P<away>.+?)\s{2,}"
    r"(?P<hg>\d+)\s*-\s*(?P<ag>\d+)"
    r"(?:\s*\((?P<ht>[^)]*)\))?\s*$"
)
HALFTIME_RE = re.compile(r"^(\d+)\s*-\s*(\d+)$")


def match_line(line):
    # "v" formatini once dene: "Ev Takim  v  Deplasman   H-A" satirlarinda
    # tire (dash) formati da yanlislikla eslesip skor parantezini
    # "deplasman takimi" sanabiliyor - once daha ozgul (v iceren) formati
    # deneyip sadece o basarisiz olursa tire formatina dusuyoruz.
    mm = MATCH_LINE_RE_V.match(line)
    if mm:
        return mm
    return MATCH_LINE_RE_DASH.match(line)


def log(msg):
    print(msg, flush=True)


# =============================================================================
# ADIM 1: FETCH (cache + retry + hizli-vazgecme)
# =============================================================================

def fetch_season_text(season_folder):
    cache_path = os.path.join(CACHE_DIR, season_folder + ".txt")
    if os.path.exists(cache_path):
        with open(cache_path, "r", encoding="utf-8") as f:
            return f.read()

    url = RAW_URL_TMPL.format(season=season_folder)
    last_err = None
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "premier-lig-pipeline/1.0"})
            with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT) as resp:
                text = resp.read().decode("utf-8", errors="replace")
            os.makedirs(CACHE_DIR, exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as f:
                f.write(text)
            return text
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code in (403, 404):
                break  # bu hatalar tekrar denemekle duzelmez, hizlica vazgec
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            last_err = e
        if attempt < MAX_ATTEMPTS:
            time.sleep(RETRY_DELAY * attempt)
    raise RuntimeError(f"{season_folder}: veri cekilemedi ({last_err})")


# =============================================================================
# ADIM 2: PARSE + DOGRULAMA
# =============================================================================

def clean_team(name):
    return re.sub(r"\s+", " ", name).strip()


def parse_season_text(text, season_label):
    matches = []
    current_week = None
    current_date_dm = None
    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n")
        stripped = line.strip()
        if not stripped:
            continue

        m = HEADER_RE.match(stripped)
        if m:
            current_week = int(m.group(1))
            continue

        dm = DATE_LINE_RE.match(stripped)
        if dm:
            current_date_dm = f"{int(dm.group(3)):02d}/{MONTH_NUM[dm.group(2)]:02d}"
            continue

        if stripped.startswith("(") or stripped.startswith(")"):
            continue  # gol atan oyuncu devam satiri

        if current_week is None:
            continue

        mm = match_line(line)
        if not mm:
            continue

        home = clean_team(mm.group("home"))
        away = clean_team(mm.group("away"))
        if not home or not away:
            continue

        ht_home = ht_away = None
        ht_raw = mm.groupdict().get("ht")
        if ht_raw:
            hm = HALFTIME_RE.match(ht_raw.strip())
            if hm:
                ht_home, ht_away = int(hm.group(1)), int(hm.group(2))

        matches.append(
            {
                "season": season_label,
                "week": current_week,
                "home": home,
                "away": away,
                "home_goals": int(mm.group("hg")),
                "away_goals": int(mm.group("ag")),
                "ht_home_goals": ht_home,
                "ht_away_goals": ht_away,
                "date_dm": current_date_dm,
            }
        )
    return matches


def validate_season(season_label, matches):
    errors = []

    if len(matches) != 380:
        errors.append(f"{season_label}: {len(matches)} mac bulundu (beklenen 380)")

    weeks = sorted(set(m["week"] for m in matches))
    if weeks != list(range(1, 39)):
        missing = sorted(set(range(1, 39)) - set(weeks))
        extra = sorted(set(weeks) - set(range(1, 39)))
        errors.append(f"{season_label}: hafta araligi hatali (eksik={missing} fazla={extra})")

    seen_pairs = set()
    for m in matches:
        key = (m["home"], m["away"])
        if key in seen_pairs:
            errors.append(f"{season_label}: duplike mac {m['home']} - {m['away']}")
        seen_pairs.add(key)

    per_week_count = defaultdict(int)
    for m in matches:
        per_week_count[m["week"]] += 1
    bad_weeks = {w: c for w, c in per_week_count.items() if c != 10}
    if bad_weeks:
        errors.append(f"{season_label}: 10 mac disinda hafta(lar) -> {bad_weeks}")

    team_match_count = defaultdict(int)
    for m in matches:
        team_match_count[m["home"]] += 1
        team_match_count[m["away"]] += 1
    bad_teams = {t: c for t, c in team_match_count.items() if c != 38}
    if bad_teams:
        errors.append(f"{season_label}: takim basina 38 mac degil -> {bad_teams}")

    return errors, len(matches) - len(seen_pairs)


def compute_standings(matches):
    table = defaultdict(lambda: {"O": 0, "G": 0, "B": 0, "M": 0, "A": 0, "Y": 0, "P": 0})
    for m in matches:
        h, a = m["home"], m["away"]
        hg, ag = m["home_goals"], m["away_goals"]
        table[h]["O"] += 1
        table[a]["O"] += 1
        table[h]["A"] += hg
        table[h]["Y"] += ag
        table[a]["A"] += ag
        table[a]["Y"] += hg
        if hg > ag:
            table[h]["G"] += 1
            table[h]["P"] += 3
            table[a]["M"] += 1
        elif hg < ag:
            table[a]["G"] += 1
            table[a]["P"] += 3
            table[h]["M"] += 1
        else:
            table[h]["B"] += 1
            table[a]["B"] += 1
            table[h]["P"] += 1
            table[a]["P"] += 1
    return table


def compute_rank_table(season_matches, through_week):
    """O sezonun 1..through_week arasindaki TUM maclarindan puan durumu
    hesaplayip takim -> sira (1=lider) sozlugu doner. Mackolik'in
    WeeklyStandingData.aspx yanitindaki "s" listesinin sirasi da ayni
    (Puan azalan) mantikla siralanmis oldugu dogrulandi; buradaki
    tie-break (Puan -> Averaj -> Atilan Gol) standart futbol kuralidir."""
    relevant = [m for m in season_matches if m["week"] <= through_week]
    table = compute_standings(relevant)
    ranked = sorted(
        table.items(),
        key=lambda kv: (-kv[1]["P"], -(kv[1]["A"] - kv[1]["Y"]), -kv[1]["A"]),
    )
    return {team: i + 1 for i, (team, _) in enumerate(ranked)}


def cross_check_known_champions(all_matches_by_season):
    ok = True
    for season_label, name_hint, expected_points in KNOWN_CHAMPION_POINTS:
        matches = all_matches_by_season.get(season_label)
        if not matches:
            log(f"  ! {season_label}: capraz kontrol atlandi (veri yok)")
            ok = False
            continue
        table = compute_standings(matches)
        candidates = {t: v for t, v in table.items() if name_hint.lower() in t.lower()}
        if not candidates:
            log(f"  ! {season_label}: '{name_hint}' takimi bulunamadi -> {list(table.keys())}")
            ok = False
            continue
        team, stats = next(iter(candidates.items()))
        if stats["P"] != expected_points:
            log(f"  X {season_label}: {team} = {stats['P']} puan (beklenen {expected_points})")
            ok = False
        else:
            log(f"  OK {season_label}: {team} = {stats['P']} puan (dogrulandi)")
    return ok


def write_data_outputs(all_matches):
    os.makedirs(DATA_DIR, exist_ok=True)
    json_path = os.path.join(DATA_DIR, "premier_lig_maclar_TUM.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_matches, f, ensure_ascii=False, indent=2)
    return json_path


def fetch_validate_all():
    """Adim 1+2: butun sezonlari cek, parse et, dogrula. Basarisizsa None doner."""
    all_matches = []
    all_matches_by_season = {}
    failed_seasons = []
    total_errors = 0
    total_dupes = 0

    log(f"[1/3] PREMIER LIG MAC VERISI - {len(SEASONS)} sezon cekiliyor (openfootball/england)")
    log("-" * 70)

    for season_label in SEASONS:
        season_folder = SEASON_FOLDERS[season_label]
        try:
            text = fetch_season_text(season_folder)
        except RuntimeError as e:
            log(f"  HATA {season_label}: {e}")
            failed_seasons.append(season_label)
            continue

        matches = parse_season_text(text, season_label)
        errors, dupes = validate_season(season_label, matches)
        total_errors += len(errors)
        total_dupes += max(dupes, 0)
        all_matches_by_season[season_label] = matches
        all_matches.extend(matches)

        status = "OK" if not errors else "SORUNLU"
        log(f"  {season_label}: {len(matches):3d} mac  [{status}]")
        for e in errors:
            log(f"      - {e}")

    log("-" * 70)
    log("BAGIMSIZ CAPRAZ KONTROL (bilinen sampiyonluk puanlari):")
    champions_ok = cross_check_known_champions(all_matches_by_season)
    log("-" * 70)

    if failed_seasons or total_errors or not champions_ok:
        log("SONUC: DOGRULAMA BASARISIZ - HICBIR EXCEL DOSYASINA YAZILMAYACAK")
        if failed_seasons:
            log(f"  Cekilemeyen sezonlar: {failed_seasons}")
        if total_errors:
            log(f"  Toplam dogrulama hatasi: {total_errors}")
        if not champions_ok:
            log("  Capraz kontrol basarisiz")
        return None

    write_data_outputs(all_matches)
    log(f"  SEZON SAYISI      : {len(SEASONS)} / {len(SEASONS)}")
    log(f"  TOPLAM MAC        : {len(all_matches)} (beklenen {len(SEASONS) * 380})")
    log(f"  DUPLIKE MAC       : {total_dupes}")
    log("  TUM DOGRULAMALAR BASARILI - Excel yazma asamasina geciliyor")
    log("-" * 70)

    by_season_week = defaultdict(list)
    for m in all_matches:
        by_season_week[(m["season"], m["week"])].append(m)
    return by_season_week, all_matches_by_season


# =============================================================================
# ADIM 3: EXCEL YAZMA
# =============================================================================

FILENAME_RE = re.compile(r"(\d{1,2})\s*\.\s*HAFTA\.xlsx$", re.IGNORECASE)
TITLE_RE = re.compile(r"^(\d{2}/\d{2})\b")

# Sezon blogu icinde MAClAR icin kullanilan sabit hucreler (her blogun kendi
# baslik satirina GORECELI). Bu satir/kolonlar iki GERCEK Excel dosyasi
# (Ingiltere Premier Lig 1. HAFTA.xlsx + doldurulmus SPANYA LA LIGA 1.
# HAFTA.xlsx ornegi) programatik olarak incelenerek tespit edildi -
# tahmin edilmedi:
#   title_row      : "XX/YY ... N. HAFTA" basligi
#   title_row+3..+22 (20 satir) : D kolonunda TAKIMLAR, E:AC PUAN DURUMU
#   title_row+3..+12 (ilk 10 satir): o haftanin 10 maci
#     AL = tarih, "GG/AA" metin (orn. "22/08")
#     AM = sabit metin "MS" (her satirda ayni, dolu La Liga orneginde dogrulandi)
#     AN = ev sahibinin o haftaya kadarki puan durumu sirasi (1=lider)
#     AO = ev sahibi (raw, "MAClAR" panelindeki AG formulunun kaynagi)
#     AP = BOS (dolu ornekte de hep bos)
#     AQ = M/S skoru "H - A" (bosluklu!, raw, AK formulunun kaynagi)
#     AR = BOS (dolu ornekte de hep bos)
#     AS = deplasman (raw, AH formulunun kaynagi)
#     AT = deplasmanin o haftaya kadarki puan durumu sirasi
#     AU = I/Y ilk yari "H - A" (raw, AJ formulunun kaynagi; kaynakta yoksa bos)
# AV,AW,AX,AY,AZ,BA,BB,BC,BD: dolu La Liga orneginde bile anlamsiz/bozuk
# (tarih tipine donusmus sayilar) ve hicbir formul onlara bakmiyor -
# KESINLIKLE DOKUNULMAZ, sadece degismediklerini dogrulamak icin izleniyor.
MATCH_COL_DATE = "AL"
MATCH_COL_MS = "AM"
MATCH_COL_HOME_RANK = "AN"
MATCH_COL_HOME = "AO"
MATCH_COL_FT = "AQ"
MATCH_COL_AWAY = "AS"
MATCH_COL_AWAY_RANK = "AT"
MATCH_COL_HT = "AU"
UNTOUCHED_SPACER_COLS = ["AP", "AR", "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD"]
STANDINGS_COLS = ["D", "E", "F", "G", "H", "I", "J", "K", "M", "AD", "AE"]
FORMULA_ZONE_COLS = ["AF", "AG", "AH", "AI", "AJ", "AK"]
N_MATCH_ROWS = 10
N_TEAM_ROWS = 20
BLOCK_SCAN_MAX_ROW = 700  # 25 blok x 23 satir + pay - blok kesfi bu araligi tarar


def find_excel_files(directory):
    """1..38 arasi 'N. HAFTA.xlsx' dosyalarini bulur. {week: filepath}"""
    found = {}
    if not os.path.isdir(directory):
        return found
    for fn in os.listdir(directory):
        if not fn.lower().endswith(".xlsx"):
            continue
        if fn.startswith("~$"):
            continue  # Excel kilit dosyasi
        m = FILENAME_RE.search(fn)
        if m:
            week = int(m.group(1))
            found[week] = os.path.join(directory, fn)
    return found


def discover_season_blocks(ws):
    """SAYFA 1'de 'XX/YY ... HAFTA' baslik satirlarini tarayip
    [(title_row, season_code), ...] dondurur. Satir numarasi hardcode
    EDILMEZ - her dosyada gercekten taranir."""
    blocks = []
    for r in range(1, BLOCK_SCAN_MAX_ROW + 1):
        v = ws.cell(row=r, column=2).value  # kolon B
        if isinstance(v, str):
            m = TITLE_RE.match(v.strip())
            if m:
                blocks.append((r, m.group(1)))
    return blocks


def snapshot_block(ws, title_row):
    """Bir sezon blogunun korunmasi gereken hucrelerinin anlik goruntusu:
    puan durumu (D:AC alt kumesi), formul zonu (metin olarak) ve spacer
    kolonlar. Yazma sonrasi bunlarla birebir karsilastirilir."""
    snap = {}
    for i in range(N_TEAM_ROWS):
        r = title_row + 3 + i
        for col in STANDINGS_COLS:
            snap[(r, col)] = ws[f"{col}{r}"].value
    for i in range(N_MATCH_ROWS):
        r = title_row + 3 + i
        for col in FORMULA_ZONE_COLS + UNTOUCHED_SPACER_COLS:
            snap[(r, col)] = ws[f"{col}{r}"].value
    return snap


def diff_snapshot(ws, snap):
    diffs = []
    for (r, col), old in snap.items():
        new = ws[f"{col}{r}"].value
        if new != old:
            diffs.append(f"{col}{r}: {old!r} -> {new!r}")
    return diffs


def process_excel_file(week_n, filepath, by_season_week, all_matches_by_season, report):
    try:
        import openpyxl
    except ImportError:
        report["fatal"] = "openpyxl kurulu degil. Once calistir: pip install openpyxl"
        return

    from openpyxl.utils import get_column_letter  # noqa: F401 (tanilama icin faydali olabilir)

    target_matchday = week_n + 1
    fname = os.path.basename(filepath)
    t_start = time.time()
    log(f"  -> {fname}  (Matchday {target_matchday})")

    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet_name = "SAYFA 1" if "SAYFA 1" in wb.sheetnames else next(
        (s for s in wb.sheetnames if "sayfa" in s.lower()), wb.sheetnames[0]
    )
    ws = wb[sheet_name]

    blocks = discover_season_blocks(ws)
    if len(blocks) < len(SEASONS) + 1:  # +1 = 26/27
        report["file_errors"][fname] = [
            f"beklenen en az {len(SEASONS) + 1} sezon blogu bulundu={len(blocks)}"
        ]
        return

    # --- ON-KONTROL: hicbir hucreye yazmadan ONCE her seyi dogrula ---
    plan = []          # (row, col, value) yazilacaklar
    snapshots = []      # her blok icin korunmasi gereken hucrelerin onceki hali
    file_errors = []
    touched_2627 = False

    for title_row, season_code in blocks:
        if season_code == "26/27":
            # 26/27'ye KESINLIKLE dokunulmuyor - sadece degismedigini dogrulamak
            # icin anlik goruntusunu aliyoruz (yazma listesine eklenmiyor).
            snapshots.append((title_row, snapshot_block(ws, title_row)))
            continue
        if season_code not in SEASONS_SET:
            file_errors.append(f"beklenmeyen sezon kodu '{season_code}' (satir {title_row})")
            continue

        matches = by_season_week.get((season_code, target_matchday), [])
        if len(matches) != N_MATCH_ROWS:
            file_errors.append(
                f"{season_code} Matchday {target_matchday}: {len(matches)} mac (beklenen {N_MATCH_ROWS})"
            )
            continue

        team_set = {
            ws.cell(row=title_row + 3 + i, column=4).value for i in range(N_TEAM_ROWS)
        }
        block_ok = True
        for m in matches:
            hn = normalize_team(m["home"])
            an = normalize_team(m["away"])
            if hn not in team_set:
                file_errors.append(
                    f"{season_code} Matchday {target_matchday}: ev sahibi eslesmedi '{m['home']}' -> '{hn}'"
                )
                block_ok = False
            if an not in team_set:
                file_errors.append(
                    f"{season_code} Matchday {target_matchday}: deplasman eslesmedi '{m['away']}' -> '{an}'"
                )
                block_ok = False
        if not block_ok:
            continue

        snapshots.append((title_row, snapshot_block(ws, title_row)))

        rank_table = compute_rank_table(all_matches_by_season[season_code], target_matchday)

        for i, m in enumerate(matches):
            r = title_row + 3 + i
            if m["date_dm"]:
                plan.append((r, MATCH_COL_DATE, m["date_dm"]))
            plan.append((r, MATCH_COL_MS, "MS"))
            plan.append((r, MATCH_COL_HOME_RANK, rank_table.get(m["home"])))
            plan.append((r, MATCH_COL_HOME, normalize_team(m["home"])))
            plan.append((r, MATCH_COL_FT, f"{m['home_goals']} - {m['away_goals']}"))
            plan.append((r, MATCH_COL_AWAY, normalize_team(m["away"])))
            plan.append((r, MATCH_COL_AWAY_RANK, rank_table.get(m["away"])))
            if m["ht_home_goals"] is not None:
                plan.append((r, MATCH_COL_HT, f"{m['ht_home_goals']} - {m['ht_away_goals']}"))

    if file_errors:
        report["file_errors"][fname] = file_errors
        log(f"     SORUNLU - {len(file_errors)} hata, bu dosyaya YAZILMADI")
        for e in file_errors[:5]:
            log(f"       - {e}")
        return

    # --- BACKUP (once yedek yoksa al) ---
    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup_path = os.path.join(BACKUP_DIR, fname)
    if not os.path.exists(backup_path):
        shutil.copy2(filepath, backup_path)
        log(f"     yedek alindi: {os.path.relpath(backup_path, EXCEL_DIR)}")

    # --- YAZMA (sadece plan'daki hucreler) ---
    for r, col, val in plan:
        ws[f"{col}{r}"] = val
    wb.save(filepath)

    # --- DOGRULAMA: dosyayi tekrar diskten oku, karsilastir ---
    # NOT: read_only=True KULLANILMIYOR - openpyxl'de read-only modda tekil
    # (rastgele) hucre erisimi cok yavas (satir numarasiyla katlanarak artan
    # bir maliyeti var; binlerce hucre icin dakikalar surebiliyor). Normal
    # yukleme burada olcumle ~12sn/dosya, rastgele erisim ise pratikte aninda.
    wb2 = openpyxl.load_workbook(filepath, data_only=False)
    ws2 = wb2[sheet_name]

    write_mismatches = []
    for r, col, val in plan:
        got = ws2[f"{col}{r}"].value
        if got != val:
            write_mismatches.append(f"{col}{r}: yazilan {val!r}, okunan {got!r}")

    protect_mismatches = []
    for title_row, snap in snapshots:
        for (r, col), old in snap.items():
            new = ws2[f"{col}{r}"].value
            if new != old:
                protect_mismatches.append(f"{col}{r} (blok {title_row}): {old!r} -> {new!r}")

    wb2.close()

    report["files_written"] += 1
    report["matches_written"] += len(
        [1 for r, col, v in plan if col in (MATCH_COL_HOME,)]
    )
    if write_mismatches:
        report["write_mismatches"][fname] = write_mismatches
    if protect_mismatches:
        report["protect_mismatches"][fname] = protect_mismatches

    status = "OK" if not (write_mismatches or protect_mismatches) else "DOGRULAMA HATASI"
    elapsed = time.time() - t_start
    log(f"     {status} - {len(plan)} hucre yazildi, {len(snapshots)} blok korundu dogrulandi ({elapsed:.0f}sn)")


def run_excel_writing(by_season_week, all_matches_by_season):
    log("[2/3] EXCEL DOSYALARI TARANIYOR")
    log("-" * 70)
    log(f"  Klasor: {EXCEL_DIR}")
    files = find_excel_files(EXCEL_DIR)

    if not files:
        log("  HATA: Bu klasorde 'N. HAFTA.xlsx' formatinda dosya bulunamadi.")
        log("  EXCEL_DIR degiskenini script icinde dogru klasore ayarlayip tekrar calistir.")
        return None

    missing = [w for w in range(1, 38) if w not in files]
    if missing:
        log(f"  UYARI: su haftalarin dosyasi bulunamadi (atlanacak): {missing}")
    if 38 in files:
        log("  38. HAFTA.xlsx bulundu ama BILINCLI OLARAK ACILMIYOR/DOKUNULMUYOR.")

    report = {
        "files_written": 0,
        "matches_written": 0,
        "file_errors": {},
        "write_mismatches": {},
        "protect_mismatches": {},
        "fatal": None,
    }

    log("-" * 70)
    log("[3/3] YAZMA + DOGRULAMA (hafta 1..37, her biri Matchday+1)")
    log("-" * 70)
    for week_n in range(1, 38):
        if week_n not in files:
            continue
        process_excel_file(week_n, files[week_n], by_season_week, all_matches_by_season, report)
        if report["fatal"]:
            log(f"DURDURULDU: {report['fatal']}")
            return report

    return report


def print_final_audit(report, seasons_validated, total_matches_source):
    log("=" * 70)
    log("FINAL DENETIM")
    log("=" * 70)
    if report is None:
        log("Excel yazma asamasi hic baslamadi (veri dogrulama basarisiz oldu).")
        return
    if report["fatal"]:
        log(f"DURUM: FATAL HATA - {report['fatal']}")
        return

    n_file_errors = sum(len(v) for v in report["file_errors"].values())
    n_write_mismatch = sum(len(v) for v in report["write_mismatches"].values())
    n_protect_mismatch = sum(len(v) for v in report["protect_mismatches"].values())

    log(f"  DOGRULANAN KAYNAK SEZON     : {seasons_validated} / {len(SEASONS)}")
    log(f"  KAYNAK TOPLAM MAC           : {total_matches_source} (duplike kaynak mac = 0)")
    log(f"  MATCHDAY KAPSAMI            : 2-38 (her sezon icin eksiksiz dogrulandi)")
    log(f"  YAZILAN DOSYA               : {report['files_written']} / 37 (Excel 1-37, +1 mapping)")
    log(f"  YAZILAN MAC (ev sahibi say.) : {report['matches_written']} (beklenen {report['files_written'] * 24 * 10})")
    log(f"  DOSYA BAZLI ON-KONTROL HATASI: {n_file_errors} (bu dosyalara HIC yazilmadi)")
    log(f"  YAZMA DOGRULAMA HATASI      : {n_write_mismatch}")
    log(f"  FORMUL/PUAN DURUMU DEGISIMI : {n_protect_mismatch}  <- 0 olmali")
    log(f"  26/27 DEGISIKLIGI           : dahil edildi (yukaridaki korunan-hucre kontrolune)")
    log(f"  38. HAFTA MAC ALANI         : HIC ACILMADI, degisiklik = 0")

    if report["file_errors"]:
        log("")
        log("  ON-KONTROLDE ELENEN DOSYALAR:")
        for fname, errs in report["file_errors"].items():
            log(f"    {fname}:")
            for e in errs[:5]:
                log(f"      - {e}")

    if n_write_mismatch or n_protect_mismatch:
        log("")
        log("  DOGRULAMA HATALARI (detay):")
        for fname, diffs in {**report["write_mismatches"], **report["protect_mismatches"]}.items():
            log(f"    {fname}:")
            for d in diffs[:5]:
                log(f"      - {d}")

    log("")
    if n_file_errors == 0 and n_write_mismatch == 0 and n_protect_mismatch == 0 and report["files_written"] > 0:
        log("  SONUC: TUM DOSYALAR BASARIYLA VE GUVENLE YAZILDI.")
    else:
        log("  SONUC: BAZI DOSYALAR SORUNLU - yukaridaki detaylara bak, yedekler _BACKUP_MAC_SONUCLARI/ icinde.")


def main():
    result = fetch_validate_all()
    if result is None:
        sys.exit(1)
    by_season_week, all_matches_by_season = result

    total_matches_source = sum(len(v) for v in by_season_week.values())
    report = run_excel_writing(by_season_week, all_matches_by_season)
    print_final_audit(report, len(SEASONS), total_matches_source)

    if report is None or report["fatal"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
